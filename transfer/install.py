import frappe
import os
from openpyxl import load_workbook
from frappe import _
import frappe
import os
from openpyxl import load_workbook


def import_branches_from_excel(filename="branch.xlsx"):
    """Import branches from Excel file (only branch names)"""

    file_path = frappe.get_app_path("transfer", "public", "asset", "xlsx", filename)

    if not os.path.exists(file_path):
        print(f"Branch file not found, skipping: {filename}")
        return False

    print(f"Importing branches from: {filename}")

    workbook = load_workbook(file_path)
    sheet = workbook.active

    imported_count = 0
    skipped_count = 0

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        # Skip empty rows
        if not any(row) or not row[0]:  # row[0] is branch name
            continue

        try:
            branch_name = str(row[0]).strip()

            # Check if branch already exists
            if not frappe.db.exists("Branch", {"branch": branch_name}):
                # Create the branch with just the name
                branch_doc = frappe.get_doc(
                    {"doctype": "Branch", "branch": branch_name}
                )
                branch_doc.insert(ignore_permissions=True)
                imported_count += 1
                print(f"✓ Created branch: {branch_name}")
            else:
                skipped_count += 1
                print(f"✓ Branch already exists: {branch_name}")

        except Exception as e:
            print(f"✗ Error processing row {row_idx}: {str(e)}")
            frappe.log_error(f"Branch import error row {row_idx}: {str(e)}")

    frappe.db.commit()
    print(
        f"Branch import completed: {imported_count} imported, {skipped_count} skipped"
    )
    return True


def check_and_import_branches():
    """Check if branches exist, import if needed"""

    # Check if any branches exist in the system
    existing_branches = frappe.get_all("Branch", limit=1)

    if existing_branches:
        print("Branches already exist in the system. Skipping import.")
        return True

    print("No branches found. Importing from Excel...")
    return import_branches_from_excel("branches.xlsx")


def check_and_set_branch_dimension():
    """Check if Branch accounting dimension is set, otherwise configure it"""

    # Check if Branch dimension exists
    branch_dimension = frappe.db.exists("Accounting Dimension", "Branch")

    if not branch_dimension:
        print("Branch accounting dimension not found. Creating it...")
        create_branch_dimension()
    else:
        print("Branch accounting dimension already exists.")
        # Verify it's properly configured
        verify_branch_dimension()


def create_branch_dimension():
    """Create Branch accounting dimension"""
    try:
        dimension = frappe.get_doc(
            {
                "doctype": "Accounting Dimension",
                "document_type": "Branch",
                "dimension_name": "Branch",
                "disabled": 0,
                "company": "",  # Applies to all companies
            }
        )
        dimension.insert(ignore_permissions=True)
        frappe.db.commit()
        print("✓ Branch accounting dimension created successfully")

        # Enable dimension in Company settings

    except Exception as e:
        print(f"✗ Failed to create Branch dimension: {str(e)}")
        frappe.log_error(f"Failed to create Branch dimension: {str(e)}")


def verify_branch_dimension():
    """Verify Branch dimension is properly configured"""
    dimension = frappe.get_doc("Accounting Dimension", "Branch")

    if dimension.disabled:
        print("Branch dimension is disabled. Enabling it...")
        dimension.disabled = 0
        dimension.save(ignore_permissions=True)
        frappe.db.commit()
        print("✓ Branch dimension enabled")

    # Check if dimension is enabled in companies

def after_install():
    print("Starting account import...")
    # Ensure Branch dimension is set up
    import_branches_from_excel("branch.xlsx")
    check_and_set_branch_dimension()
    # Get the default company
    default_company = frappe.db.get_single_value("Global Defaults", "default_company")
    if not default_company:
        default_company = frappe.db.get_value("Company", {}, "name")

    if not default_company:
        frappe.throw("No Company found. Please create a Company first.")

    print(f"Using company: {default_company}")

    # Import accounts
    import_accounts_from_excel(default_company, "roots_accounts.xlsx")
    import_accounts_from_excel(default_company, "accounts.xlsx")
    print("Account import completed successfully!")


def import_accounts_from_excel(company, filename):
    """Import accounts from Excel file without relying on autoname"""

    file_path = frappe.get_app_path("transfer", "public", "asset", "xlsx", filename)

    if not os.path.exists(file_path):
        print(f"File not found, skipping: {filename}")
        return

    print(f"Importing accounts from: {filename}")

    workbook = load_workbook(file_path)
    sheet = workbook.active
    abbr = frappe.db.get_value("Company", company, "abbr") or "CO"

    accounts = []

    # Read accounts from Excel
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        # Skip empty rows
        if not any(row) or not row[1]:  # row[1] is account_name
            continue

        try:
            account_data = {
                "doctype": "Account",
                "account_name": str(row[1]).strip(),
                "account_number": str(row[2]).strip() if row[2] else "",
                "is_group": bool(int(row[3])) if row[3] is not None else False,
                "account_currency": str(row[4]).strip() if row[4] else "LYD",
                "company": company,
                "report_type": "Balance Sheet",
                "root_type": str(row[5]).strip() if row[5] else "",
                "account_type": str(row[6]).strip() if row[6] else "Asset",
            }

            # Set parent account for non-group accounts
            if not account_data["is_group"] and row[5]:
                account_data["parent_account"] = str(row[5]).strip()

            # MANUALLY set the account name to avoid autoname issues
            if account_data["account_number"]:
                account_data["name"] = (
                    f"{account_data['account_number']} - {account_data['account_name']} - {abbr}"
                )
            else:
                account_data["name"] = f"{account_data['account_name']} - {abbr}"

            accounts.append(account_data)

        except Exception as e:
            print(f"Error processing row {row_idx}: {e}")
            continue

    # Import accounts in order (groups first)
    group_accounts = [acc for acc in accounts if acc["is_group"]]
    child_accounts = [acc for acc in accounts if not acc["is_group"]]

    # Import group accounts first
    for account in group_accounts:
        create_account(account)

    # Import child accounts
    for account in child_accounts:
        create_account(account)


def create_account(account_data):
    """Create a single account with proper error handling"""

    try:
        # Check if account already exists
        if frappe.db.exists("Account", account_data["name"]):
            print(f"✓ Account already exists: {account_data['name']}")
            return

        # Create the account
        doc = frappe.get_doc(account_data)
        doc.insert(
            ignore_permissions=True,
            ignore_links=True,
            ignore_mandatory=True,
            ignore_if_duplicate=True,
        )

        print(f"✓ Created account: {account_data['name']}")
        frappe.db.commit()

    except frappe.DuplicateEntryError:
        print(f"✓ Account already exists (duplicate): {account_data['name']}")
    except Exception as e:
        print(f"✗ Failed to create account {account_data['name']}: {str(e)}")
        frappe.log_error(f"Account creation failed: {str(e)}")


def after_uninstall():
    """Clean up accounts on uninstall"""
    print("Uninstalling transfer app accounts...")

    # Get accounts to delete from both files
    accounts_to_delete = []

    for filename in ["roots_accounts.xlsx", "accounts.xlsx"]:
        file_path = frappe.get_app_path("transfer", "public", "asset", "xlsx", filename)

        if os.path.exists(file_path):
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[1]:  # account_name
                        accounts_to_delete.append(str(row[1]).strip())

            except Exception as e:
                print(f"Error reading {filename}: {e}")

    # Delete accounts
    for account_name in accounts_to_delete:
        try:
            if frappe.db.exists("Account", {"account_name": account_name}):
                frappe.delete_doc("Account", account_name, ignore_permissions=True)
                print(f"Deleted account: {account_name}")
        except Exception as e:
            print(f"Failed to delete account {account_name}: {e}")

    print("removing accounts completed")
